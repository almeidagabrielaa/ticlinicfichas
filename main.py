import os
import glob
import demografia
import atendimentos
import epidemiologia
import usuarios
import perfil_adoecimento
import resultados_exames
import resultados_encaminhamentos_prescricoes
import import_excel_to_sqlite
from openpyxl import load_workbook

def main():
    # Chama funções principais de cada módulo
    demografia.analise_demografia()
    atendimentos.analise_atendimentos()
    epidemiologia.analise_epidemiologia()
    usuarios.analise_usuarios()
    perfil_adoecimento.analise_perfil_adoecimento()
    resultados_exames.analise_resultados_exames()
    resultados_encaminhamentos_prescricoes.analise_resultados_encaminhamentos_prescricoes()
    
    # Verificar se o banco de dados existe ou criar
    db_file = 'fichas_maceio.db'
    if not os.path.exists(db_file):
        open(db_file, 'a').close()  # Cria o arquivo se não existir

    # Busca por todos os arquivos .xlsx na raiz do projeto
    for excel_file in glob.glob("*.xlsx"):
        print(f"Processando o arquivo: {excel_file}")

        # Usar o nome do arquivo como nome da tabela, removendo a extensão
        table_name = os.path.splitext(excel_file)[0]

        # Obter o nome da primeira aba
        workbook = load_workbook(filename=excel_file, read_only=True)
        sheet_name = workbook.sheetnames[0]

        # Importa dados de cada arquivo Excel para o SQLite
        import_excel_to_sqlite.import_excel_to_sqlite(
            excel_file_path=excel_file,
            sheet_name=sheet_name,
            db_file=db_file,
            table_name=table_name
        )

if __name__ == "__main__":
    main()